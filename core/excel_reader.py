"""Excel 解析器 — 读取 .xlsx 文件，按 Sheet 返回 List[Dict]。

行过滤、类型规范化等清洗逻辑统一委托给 DataCleaner。
"""
from pathlib import Path
from typing import Any, Dict, List, Optional

from loguru import logger
from openpyxl import load_workbook

from core.file_reader import DataCleaner, parse_json_field


class ExcelReader:
    """读取 .xlsx 文件，支持：
    - 多 Sheet 读取
    - 行过滤（跳过注释行、空行、skip=Y）
    - 自动 JSON 字符串解析
    """

    @staticmethod
    def read_sheet(file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """读取指定 Sheet 的所有用例行

        Args:
            file_path: Excel 文件路径
            sheet_name: Sheet 名称，默认读取第一个 Sheet
        Returns:
            用例行列表，每行为 dict，key 为表头字段名
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {file_path}")

        wb = load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name] if sheet_name else wb.active
        result = ExcelReader._read_sheet_from_wb(wb, sheet, file_path)
        wb.close()
        return result

    @staticmethod
    def _read_sheet_from_wb(wb, sheet, file_path: str) -> List[Dict[str, Any]]:
        """从已打开的 Workbook 和 Worksheet 对象读取用例行。"""
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) < 2:
            return []

        # 第一行为表头，构建原始行数据列表（空表头生成占位名）
        headers = []
        for h in rows[0]:
            name = str(h).strip() if h else ""
            if not name:
                name = f"_col_{len(headers) + 1}"
            headers.append(name)
        raw_rows = []
        for row in rows[1:]:
            row_dict = {}
            for idx, header in enumerate(headers):
                row_dict[header] = row[idx] if idx < len(row) else None
            raw_rows.append(row_dict)

        # 统一委托 DataCleaner 做清洗（过滤、规范化、别名映射）
        cases = DataCleaner.clean_rows(raw_rows)
        logger.debug("读取 Excel: {} → Sheet: {}", file_path, sheet.title)
        return cases

    @staticmethod
    def read_all_sheets(file_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """读取所有 Sheet，返回 {sheet_name: List[Dict]}（只打开一次文件）。"""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {file_path}")

        wb = load_workbook(file_path, data_only=True)
        result = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            cases = ExcelReader._read_sheet_from_wb(wb, sheet, file_path)
            result[sheet_name] = cases
            logger.info("从 {} 读取到 {} 条用例", file_path, len(cases))
        wb.close()
        return result

    @staticmethod
    def read_directory(dir_path: str, pattern: str = "*.xlsx") -> List[Dict[str, Any]]:
        """递归读取目录下所有 Excel 文件的所有 Sheet 用例

        Args:
            dir_path: 目录路径
            pattern: 文件匹配模式，默认 *.xlsx
        Returns:
            所有用例行列表
        """
        base = Path(dir_path)
        if not base.exists():
            logger.warning("目录不存在: {}", dir_path)
            return []

        all_cases = []
        for xlsx_file in sorted(base.rglob(pattern)):
            # 跳过临时文件（以 ~$ 开头）
            if xlsx_file.name.startswith("~$"):
                continue
            try:
                sheets = ExcelReader.read_all_sheets(str(xlsx_file))
                for sheet_name, cases in sheets.items():
                    for case in cases:
                        # 注入来源信息
                        case["_source_file"] = str(xlsx_file)
                        case["_source_sheet"] = sheet_name
                    all_cases.extend(cases)
            except Exception as e:
                logger.error("读取文件失败: {} | 错误: {}", xlsx_file, e)

        logger.info("从目录 {} 共读取到 {} 条用例", dir_path, len(all_cases))
        return all_cases
