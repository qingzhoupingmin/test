"""模板生成器 — 快速生成多格式测试用例模板文件。

支持格式: xlsx (Excel)、csv、json、yaml
"""

import csv
import json
import os
from pathlib import Path
from typing import Optional


# ==================== 共用数据 ====================

# 表头/字段列表（所有格式通用）
_API_HEADERS = [
    "case_id", "case_name", "module", "tags", "priority",
    "method", "url", "headers", "params", "body",
    "payload_type", "assertions", "extract",
    "pre_hook", "post_hook", "skip", "depends_on", "retry",
]

# 示例数据
_API_SAMPLES = [
    {
        "case_id": "TC_EXAMPLE_001",
        "case_name": "获取资源列表",
        "module": "示例模块",
        "tags": "smoke,p0",
        "priority": "P0",
        "method": "GET",
        "url": "/api/v1/resources",
        "headers": "",
        "params": '{"page": 1, "size": 10}',
        "body": "",
        "payload_type": "",
        "assertions": '[{"type": "status_code", "value": 200}, {"type": "jsonpath_exists", "key": "$.data"}]',
        "extract": "$.data",
        "pre_hook": "",
        "post_hook": "",
        "skip": "否",
        "depends_on": "",
        "retry": 0,
    },
    {
        "case_id": "TC_EXAMPLE_002",
        "case_name": "创建资源",
        "module": "示例模块",
        "tags": "p0,regression",
        "priority": "P0",
        "method": "POST",
        "url": "/api/v1/resources",
        "headers": '{"Content-Type": "application/json"}',
        "params": "",
        "body": '{"name": "test_resource", "desc": "auto_test"}',
        "payload_type": "json",
        "assertions": '[{"type": "status_code", "value": 201}, {"type": "jsonpath_equals", "key": "$.data.name", "value": "test_resource"}]',
        "extract": "$.data.id",
        "pre_hook": "hooks.sample_hooks.hook_before_case",
        "post_hook": "hooks.sample_hooks.hook_after_case",
        "skip": "否",
        "depends_on": "",
        "retry": 1,
    },
    {
        "case_id": "TC_EXAMPLE_003",
        "case_name": "查询资源详情",
        "module": "示例模块",
        "tags": "p1",
        "priority": "P1",
        "method": "GET",
        "url": "/api/v1/resources/{{new_resource_id}}",
        "headers": "",
        "params": "",
        "body": "",
        "payload_type": "",
        "assertions": '[{"type": "status_code", "value": 200}]',
        "extract": "",
        "pre_hook": "",
        "post_hook": "",
        "skip": "否",
        "depends_on": "TC_EXAMPLE_002",
        "retry": 0,
    },
    {
        "case_id": "TC_EXAMPLE_004",
        "case_name": "删除资源",
        "module": "示例模块",
        "tags": "p1",
        "priority": "P1",
        "method": "DELETE",
        "url": "/api/v1/resources/{{new_resource_id}}",
        "headers": "",
        "params": "",
        "body": "",
        "payload_type": "",
        "assertions": '[{"type": "status_code", "value": 204}]',
        "extract": "",
        "pre_hook": "",
        "post_hook": "",
        "skip": "否",
        "depends_on": "TC_EXAMPLE_003",
        "retry": 0,
    },
]


class TemplateGenerator:
    """生成多格式测试用例模板文件。"""

    TEMPLATE_DIR = Path(__file__).parent

    @classmethod
    def generate_api_template(
        cls,
        output_path: Optional[str] = None,
        fmt: str = "xlsx",
    ) -> str:
        """生成 API 用例模板文件（根据 fmt 自动选择格式）。

        Args:
            output_path: 输出文件路径（None 则自动生成到 templates/ 目录）
            fmt: 输出格式，可选 "xlsx"、"csv"、"json"、"yaml"
        Returns:
            生成的文件路径
        """
        fmt = fmt.lower().strip()
        ext_map = {
            "xlsx": ".xlsx",
            "excel": ".xlsx",
            "csv": ".csv",
            "json": ".json",
            "yaml": ".yaml",
            "yml": ".yaml",
        }
        if fmt not in ext_map:
            raise ValueError(f"不支持的模板格式: {fmt}，可选: xlsx, csv, json, yaml")

        ext = ext_map[fmt]
        if output_path is None:
            output_path = str(cls.TEMPLATE_DIR / f"api_sample{ext}")

        # 确保输出目录存在
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        if fmt in ("xlsx", "excel"):
            return cls._generate_xlsx(output_path)
        elif fmt == "csv":
            return cls._generate_csv(output_path)
        elif fmt == "json":
            return cls._generate_json(output_path)
        elif fmt in ("yaml", "yml"):
            return cls._generate_yaml(output_path)

        return output_path

    # ==================== Excel 模板 ====================

    @classmethod
    def _generate_xlsx(cls, output_path: str) -> str:
        """生成 Excel 模板。"""
        openpyxl = cls._ensure_openpyxl()

        wb = openpyxl.Workbook()

        # Sheet 1: 用例列表
        ws = wb.active
        ws.title = "API用例"

        # 表头
        for col, h in enumerate(_API_HEADERS, 1):
            ws.cell(row=1, column=col, value=h)

        # 示例数据行
        for row_idx, sample in enumerate(_API_SAMPLES, 2):
            for col_idx, key in enumerate(_API_HEADERS, 1):
                ws.cell(row=row_idx, column=col_idx, value=sample.get(key, ""))

        # Sheet 2: 使用说明
        ws2 = wb.create_sheet("使用说明")
        instructions = [
            ["字段", "说明", "示例"],
            ["case_id", "用例唯一ID", "TC_USER_001"],
            ["case_name", "用例名称", "用户登录成功"],
            ["module", "所属模块", "用户管理"],
            ["tags", "标签(逗号分隔)", "smoke,p0,regression"],
            ["priority", "优先级", "P0 / P1 / P2"],
            ["method", "请求方法", "GET/POST/PUT/DELETE/PATCH"],
            ["url", "接口路径", "/api/v1/login"],
            ["headers", "请求头(JSON)", '{"Content-Type":"application/json"}'],
            ["params", "URL参数(JSON)", '{"page":1,"size":10}'],
            ["body", "请求体(JSON或字符串)", '{"name":"test"}'],
            ["payload_type", "请求体类型", "json / form / formdata"],
            ["assertions", "断言规则(JSON数组)", '[{"type":"status_code","value":200}]'],
            ["extract", "响应提取规则", 'token=$.data.token; uid=$.data.id'],
            ["pre_hook", "前置钩子", "hooks.sample_hooks.hook_before_case"],
            ["post_hook", "后置钩子", "hooks.sample_hooks.hook_after_case"],
            ["skip", "是否跳过", "是/否"],
            ["depends_on", "依赖的case_id", "TC_USER_001"],
            ["retry", "失败重试次数", "1"],
        ]
        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, val in enumerate(row_data, 1):
                ws2.cell(row=row_idx, column=col_idx, value=val)

        # 调整列宽
        for ws_obj in [ws, ws2]:
            for col in ws_obj.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                ws_obj.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

        wb.save(output_path)
        print(f"[xlsx] 模板已生成: {output_path}")
        return output_path

    # ==================== CSV 模板 ====================

    @classmethod
    def _generate_csv(cls, output_path: str) -> str:
        """生成 CSV 模板。"""
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=_API_HEADERS)
            writer.writeheader()
            for sample in _API_SAMPLES:
                writer.writerow(sample)

        print(f"[csv] 模板已生成: {output_path}")
        return output_path

    # ==================== JSON 模板 ====================

    @classmethod
    def _generate_json(cls, output_path: str) -> str:
        """生成 JSON 模板（数组格式）。"""
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(_API_SAMPLES, f, ensure_ascii=False, indent=2)

        print(f"[json] 模板已生成: {output_path}")
        return output_path

    # ==================== YAML 模板 ====================

    @classmethod
    def _generate_yaml(cls, output_path: str) -> str:
        """生成 YAML 模板（列表格式）。"""
        try:
            import yaml
        except ImportError:
            raise ImportError("生成 YAML 模板需要 PyYAML 库，请执行: pip install pyyaml")

        with open(output_path, "w", encoding="utf-8") as f:
            yaml.dump(
                _API_SAMPLES,
                f,
                allow_unicode=True,
                default_flow_style=False,
                sort_keys=False,
            )

        print(f"[yaml] 模板已生成: {output_path}")
        return output_path

    # ==================== 工具方法 ====================

    @staticmethod
    def _ensure_openpyxl():
        """确保 openpyxl 可用。"""
        try:
            import openpyxl
            return openpyxl
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")


def main():
    """生成全部格式的 API 测试用例模板。"""
    print("=" * 50)
    print("API 测试框架模板生成器")
    print("=" * 50)

    # 支持命令行指定格式: python template_generator.py [xlsx|csv|json|yaml|all]
    import sys
    formats = sys.argv[1:] if len(sys.argv) > 1 else ["all"]

    if "all" in formats:
        formats = ["xlsx", "csv", "json", "yaml"]

    for fmt in formats:
        try:
            TemplateGenerator.generate_api_template(fmt=fmt)
        except Exception as e:
            print(f"[错误] 生成 {fmt} 模板失败: {e}")

    print("模板生成完成！")


if __name__ == "__main__":
    main()